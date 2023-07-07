import requests
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

url1 = 'https://www.alphavantage.co/query?function=TIME_SERIES_INTRADAY&symbol=PYPL&interval=5min&apikey=TFK5P3Y74NWKG8U8'
url2 = 'https://www.alphavantage.co/query?function=TIME_SERIES_INTRADAY&symbol=NTLA&interval=5min&apikey=TFK5P3Y74NWKG8U8'

r1 = requests.get(url1)
r2 = requests.get(url2)

data1 = r1.json()
data2 = r2.json()


wb = Workbook()

default_sheet = wb.active
wb.remove(default_sheet)


if 'Time Series (5min)' in data1:
    time_series1 = data1['Time Series (5min)']

    
    sheet1 = wb.create_sheet(title='PYPL')
    sheet1.append(['Timestamp', 'Open', 'High', 'Low', 'Close', 'Volume'])

    
    for timestamp, values in time_series1.items():
        row = [timestamp, float(values['1. open']), float(values['2. high']), float(values['3. low']),
               float(values['4. close']), int(values['5. volume'])]
        sheet1.append(row)

    print("PYPL stock data added to the Excel file.")

else:
    print("Failed to retrieve PYPL stock data.")


if 'Time Series (5min)' in data2:
    time_series2 = data2['Time Series (5min)']

  
    sheet2 = wb.create_sheet(title='NTLA')
    sheet2.append(['Timestamp', 'Open', 'High', 'Low', 'Close', 'Volume'])

  
    for timestamp, values in time_series2.items():
        row = [timestamp, float(values['1. open']), float(values['2. high']), float(values['3. low']),
               float(values['4. close']), int(values['5. volume'])]
        sheet2.append(row)

    print("NTLA stock data added to the Excel file.")

else:
    print("Failed to retrieve NTLA stock data.")


for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    for column in sheet.columns:
        column_letter = get_column_letter(column[0].column)
        if column[0].column != 1:
            for cell in column:
                cell.number_format = '0.00'


wb.save('stock_data.xlsx')
print("Stock data saved to 'stock_data.xlsx' file.")
